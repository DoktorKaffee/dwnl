#!/usr/bin/env python3
"""
CLI-инструмент для инвентаризации серверов и СХД через Redfish/Swordfish.
Собирает информацию о всех комплектующих: диски, память, CPU, сеть, питание, вентиляторы, контроллеры.

Использование:
  python inventory_cli.py -u admin -p password -H 192.168.1.100
  python inventory_cli.py -u admin -p password -H localhost:8000 --no-ssl --disks
  python inventory_cli.py -u admin -p password -H 192.168.1.100 --all --excel report.xlsx
"""

import argparse
import json
import sys
from typing import Dict, Any, List, Optional
from urllib.parse import urljoin

try:
    import httpx
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn
except ImportError as e:
    print(f"❌ Ошибка импорта: {e}")
    print("Установите зависимости: pip install httpx rich openpyxl")
    sys.exit(1)

# Попытка импорта openpyxl для Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

console = Console()


# ============================================
# КЛИЕНТ REDFISH
# ============================================

class RedfishClient:
    """Клиент для работы с Redfish/Swordfish API"""

    def __init__(self, host: str, username: str, password: str, verify_ssl: bool = False, use_ssl: bool = True):
        """
        :param host: Адрес хоста (может включать протокол и порт, например localhost:8000 или https://192.168.1.1)
        :param username: Имя пользователя
        :param password: Пароль
        :param verify_ssl: Проверять SSL-сертификат (только для HTTPS)
        :param use_ssl: Использовать HTTPS (если True) или HTTP (если False)
        """
        # Если в host уже указан протокол, оставляем как есть, иначе добавляем на основе use_ssl
        if host.startswith(("http://", "https://")):
            self.base_url = host.rstrip('/')
        else:
            scheme = "https" if use_ssl else "http"
            self.base_url = f"{scheme}://{host}"
        self.username = username
        self.password = password
        self.verify_ssl = verify_ssl
        self.session = httpx.Client(verify=verify_ssl, timeout=30.0)

    def _request(self, method: str, path: str, **kwargs) -> httpx.Response:
        url = urljoin(self.base_url + '/', path.lstrip('/'))  # urljoin корректно объединит
        response = self.session.request(
            method,
            url,
            auth=(self.username, self.password),
            headers={"Accept": "application/json"},
            **kwargs
        )
        response.raise_for_status()
        return response

    def get(self, path: str) -> Dict[str, Any]:
        """Выполняет GET-запрос и возвращает JSON-словарь"""
        response = self._request("GET", path)
        return response.json()

    def follow_collection(self, collection_path: str) -> List[Dict[str, Any]]:
        """
        Получает все элементы коллекции, следуя пагинации.
        Redfish использует odata.nextLink для постраничной загрузки.
        """
        items = []
        current_path = collection_path
        while current_path:
            data = self.get(current_path)
            members = data.get("Members", [])
            for member in members:
                # member может быть ссылкой (dict с @odata.id) или объектом
                if isinstance(member, dict) and "@odata.id" in member:
                    try:
                        member_data = self.get(member["@odata.id"])
                        items.append(member_data)
                    except Exception:
                        # Если не удалось получить детали, добавляем только ссылку
                        items.append(member)
                else:
                    items.append(member)
            # Пагинация
            next_link = data.get("@odata.nextLink")
            current_path = next_link if next_link else None
        return items


# ============================================
# СБОР ДАННЫХ
# ============================================

def collect_inventory(client: RedfishClient) -> Dict[str, List[Dict[str, Any]]]:
    """Собирает все данные с устройства"""
    inventory = {
        "systems": [],
        "chassis": [],
        "storage": [],
        "managers": [],
        "disks": [],
        "memory": [],
        "cpu": [],
        "network": [],
        "power": [],
        "fans": [],
        "controllers": [],
        "backplanes": [],
    }

    # 1. Системы (серверы)
    try:
        systems = client.follow_collection("/redfish/v1/Systems")
        for sys in systems:
            inventory["systems"].append(sys)
            # Из системы можно извлечь CPU, память, сеть
            if "Processors" in sys and "@odata.id" in sys["Processors"]:
                cpus = client.follow_collection(sys["Processors"]["@odata.id"])
                for cpu in cpus:
                    inventory["cpu"].append(cpu)
            if "Memory" in sys and "@odata.id" in sys["Memory"]:
                memory = client.follow_collection(sys["Memory"]["@odata.id"])
                for mem in memory:
                    inventory["memory"].append(mem)
            if "EthernetInterfaces" in sys and "@odata.id" in sys["EthernetInterfaces"]:
                net = client.follow_collection(sys["EthernetInterfaces"]["@odata.id"])
                for nic in net:
                    inventory["network"].append(nic)
    except Exception as e:
        console.print(f"[yellow]⚠️ Ошибка получения Systems: {e}[/]")

    # 2. Шасси (питание, вентиляторы, диски)
    try:
        chassis = client.follow_collection("/redfish/v1/Chassis")
        for ch in chassis:
            inventory["chassis"].append(ch)
            # В шасси могут быть диски (Drives), блоки питания (PowerSupplies), вентиляторы (Fans)
            if "Drives" in ch and "@odata.id" in ch["Drives"]:
                drives = client.follow_collection(ch["Drives"]["@odata.id"])
                for disk in drives:
                    inventory["disks"].append(disk)
            if "PowerSupplies" in ch and "@odata.id" in ch["PowerSupplies"]:
                power = client.follow_collection(ch["PowerSupplies"]["@odata.id"])
                for psu in power:
                    inventory["power"].append(psu)
            if "Fans" in ch and "@odata.id" in ch["Fans"]:
                fans = client.follow_collection(ch["Fans"]["@odata.id"])
                for fan in fans:
                    inventory["fans"].append(fan)
    except Exception as e:
        console.print(f"[yellow]⚠️ Ошибка получения Chassis: {e}[/]")

    # 3. Хранилище (Storage) — диски, контроллеры, логические тома
    try:
        storage_collection = client.get("/redfish/v1/Storage")
        storages = client.follow_collection("/redfish/v1/Storage")
        for st in storages:
            inventory["storage"].append(st)
            # Диски внутри Storage
            if "Drives" in st and "@odata.id" in st["Drives"]:
                drives = client.follow_collection(st["Drives"]["@odata.id"])
                for disk in drives:
                    inventory["disks"].append(disk)
            # Контроллеры (StorageControllers)
            if "StorageControllers" in st:
                controllers = st["StorageControllers"]
                if isinstance(controllers, list):
                    inventory["controllers"].extend(controllers)
                else:
                    # может быть ссылка
                    pass
            # Backplanes? В некоторых реализациях есть отдельный ресурс
    except Exception as e:
        console.print(f"[yellow]⚠️ Ошибка получения Storage: {e}[/]")

    # 4. Managers (для сетевых интерфейсов BMC)
    try:
        managers = client.follow_collection("/redfish/v1/Managers")
        for mgr in managers:
            inventory["managers"].append(mgr)
            # NetworkInterfaces в менеджере (для BMC)
            if "NetworkInterfaces" in mgr and "@odata.id" in mgr["NetworkInterfaces"]:
                net = client.follow_collection(mgr["NetworkInterfaces"]["@odata.id"])
                for nic in net:
                    inventory["network"].append(nic)
    except Exception as e:
        console.print(f"[yellow]⚠️ Ошибка получения Managers: {e}[/]")

    # 5. Дополнительно: поиск дисков в SimpleStorage (устаревший, но может быть)
    try:
        # Некоторые системы используют SimpleStorage
        simple_storage = client.follow_collection("/redfish/v1/Systems/1/SimpleStorage")
        for ss in simple_storage:
            if "Devices" in ss:
                for dev in ss["Devices"]:
                    # преобразуем в диск
                    disk = {
                        "Name": dev.get("Name"),
                        "CapacityBytes": dev.get("SizeBytes"),
                        "Manufacturer": dev.get("Manufacturer"),
                        "Model": dev.get("Model"),
                        "SerialNumber": dev.get("SerialNumber"),
                        "Status": dev.get("Status"),
                        "Interface": dev.get("Interface"),
                        "Type": "SimpleStorage"
                    }
                    inventory["disks"].append(disk)
    except Exception:
        pass

    return inventory


# ============================================
# ПАРСИНГ ДАННЫХ В УНИФИЦИРОВАННЫЙ ФОРМАТ
# ============================================

def parse_memory(mem: Dict) -> Dict:
    """Парсит модуль памяти"""
    return {
        "Name": mem.get("Name", ""),
        "CapacityGB": mem.get("CapacityMB", 0) / 1024 if "CapacityMB" in mem else mem.get("CapacityGB", 0),
        "Type": mem.get("MemoryType", ""),
        "SpeedMHz": mem.get("SpeedMHz", ""),
        "Manufacturer": mem.get("Manufacturer", ""),
        "Serial": mem.get("SerialNumber", ""),
        "PartNumber": mem.get("PartNumber", ""),
        "Status": mem.get("Status", {}).get("Health", "OK") if "Status" in mem else "OK",
        "Location": mem.get("Location", {}).get("PartLocation", {}).get("ServiceLabel", ""),
    }

def parse_cpu(cpu: Dict) -> Dict:
    return {
        "Name": cpu.get("Name", ""),
        "Model": cpu.get("Model", ""),
        "Manufacturer": cpu.get("Manufacturer", ""),
        "Cores": cpu.get("TotalCores", 0),
        "Threads": cpu.get("TotalThreads", 0),
        "MaxSpeedMHz": cpu.get("MaxSpeedMHz", 0),
        "Socket": cpu.get("Socket", ""),
        "Status": cpu.get("Status", {}).get("Health", "OK") if "Status" in cpu else "OK",
    }

def parse_disk(disk: Dict) -> Dict:
    """Парсит диск (Drive)"""
    status = disk.get("Status", {})
    health = status.get("Health", "OK") if status else "OK"
    return {
        "Name": disk.get("Name", ""),
        "Model": disk.get("Model", ""),
        "Serial": disk.get("SerialNumber", ""),
        "CapacityGB": disk.get("CapacityBytes", 0) / (1024**3) if "CapacityBytes" in disk else 0,
        "Interface": disk.get("Interface", ""),
        "Type": disk.get("MediaType", ""),
        "RPM": disk.get("RotationSpeedRPM", 0),
        "Firmware": disk.get("FirmwareVersion", ""),
        "Status": health,
        "Manufacturer": disk.get("Manufacturer", ""),
        "PartNumber": disk.get("PartNumber", ""),
        "Location": disk.get("Location", {}).get("PartLocation", {}).get("ServiceLabel", ""),
    }

def parse_controller(ctl: Dict) -> Dict:
    status = ctl.get("Status", {})
    health = status.get("Health", "OK") if status else "OK"
    return {
        "Name": ctl.get("Name", ""),
        "Model": ctl.get("Model", ""),
        "Manufacturer": ctl.get("Manufacturer", ""),
        "Firmware": ctl.get("FirmwareVersion", ""),
        "Status": health,
        "Serial": ctl.get("SerialNumber", ""),
        "PartNumber": ctl.get("PartNumber", ""),
    }

def parse_power(psu: Dict) -> Dict:
    status = psu.get("Status", {})
    health = status.get("Health", "OK") if status else "OK"
    return {
        "Name": psu.get("Name", ""),
        "Model": psu.get("Model", ""),
        "Manufacturer": psu.get("Manufacturer", ""),
        "PowerWatts": psu.get("PowerCapacityWatts", 0),
        "InputVoltage": psu.get("InputVoltage", 0),
        "Status": health,
        "Serial": psu.get("SerialNumber", ""),
    }

def parse_fan(fan: Dict) -> Dict:
    status = fan.get("Status", {})
    health = status.get("Health", "OK") if status else "OK"
    return {
        "Name": fan.get("Name", ""),
        "Status": health,
        "SpeedRPM": fan.get("SpeedRPM", 0),
        "SpeedPercent": fan.get("SpeedPercent", 0),
    }

def parse_network(nic: Dict) -> Dict:
    status = nic.get("Status", {})
    health = status.get("Health", "OK") if status else "OK"
    return {
        "Name": nic.get("Name", ""),
        "Description": nic.get("Description", ""),
        "MAC": nic.get("MACAddress", ""),
        "SpeedMbps": nic.get("SpeedMbps", 0),
        "Status": health,
        "Manufacturer": nic.get("Manufacturer", ""),
        "Model": nic.get("Model", ""),
        "Firmware": nic.get("FirmwareVersion", ""),
    }

def parse_system(sys: Dict) -> Dict:
    return {
        "Name": sys.get("Name", ""),
        "Model": sys.get("Model", ""),
        "Manufacturer": sys.get("Manufacturer", ""),
        "Serial": sys.get("SerialNumber", ""),
        "Firmware": sys.get("FirmwareVersion", ""),
        "Status": sys.get("Status", {}).get("Health", "OK") if "Status" in sys else "OK",
        "PowerState": sys.get("PowerState", ""),
    }


# ============================================
# ВЫВОД В КОНСОЛЬ (Rich)
# ============================================

def print_table(title: str, columns: List[str], data: List[Dict], field_map: Dict[str, str]):
    """Выводит таблицу с помощью Rich"""
    if not data:
        return
    table = Table(title=title)
    for col in columns:
        table.add_column(col)
    for row in data:
        row_values = []
        for col in columns:
            key = field_map.get(col, col)
            val = row.get(key, "")
            # преобразуем в строку
            if isinstance(val, (int, float)):
                if key in ["CapacityGB", "PowerWatts", "SpeedMbps", "SpeedRPM"] and val > 0:
                    row_values.append(f"{val:.2f}" if isinstance(val, float) else str(val))
                else:
                    row_values.append(str(val))
            elif val is None:
                row_values.append("")
            else:
                row_values.append(str(val))
        table.add_row(*row_values)
    console.print(table)


def print_inventory(inventory: Dict[str, List[Dict]], selected_types: List[str]):
    """Выводит собранный инвентарь в виде таблиц"""
    # Словари для отображения полей
    field_maps = {
        "disks": {
            "Name": "Name", "Model": "Model", "Serial": "Serial", "CapacityGB": "Capacity (GB)",
            "Interface": "Interface", "Type": "Type", "RPM": "RPM", "Firmware": "Firmware",
            "Status": "Status", "Manufacturer": "Manufacturer", "PartNumber": "Part Number",
            "Location": "Location"
        },
        "memory": {
            "Name": "Name", "CapacityGB": "Capacity (GB)", "Type": "Type", "SpeedMHz": "Speed (MHz)",
            "Manufacturer": "Manufacturer", "Serial": "Serial", "PartNumber": "Part Number",
            "Status": "Status", "Location": "Location"
        },
        "cpu": {
            "Name": "Name", "Model": "Model", "Manufacturer": "Manufacturer",
            "Cores": "Cores", "Threads": "Threads", "MaxSpeedMHz": "Max Speed (MHz)",
            "Socket": "Socket", "Status": "Status"
        },
        "network": {
            "Name": "Name", "Description": "Description", "MAC": "MAC Address",
            "SpeedMbps": "Speed (Mbps)", "Status": "Status", "Manufacturer": "Manufacturer",
            "Model": "Model", "Firmware": "Firmware"
        },
        "power": {
            "Name": "Name", "Model": "Model", "Manufacturer": "Manufacturer",
            "PowerWatts": "Power (W)", "InputVoltage": "Input (V)", "Status": "Status",
            "Serial": "Serial"
        },
        "fans": {
            "Name": "Name", "Status": "Status", "SpeedRPM": "Speed (RPM)", "SpeedPercent": "Speed (%)"
        },
        "controllers": {
            "Name": "Name", "Model": "Model", "Manufacturer": "Manufacturer",
            "Firmware": "Firmware", "Status": "Status", "Serial": "Serial", "PartNumber": "Part Number"
        },
        "systems": {
            "Name": "Name", "Model": "Model", "Manufacturer": "Manufacturer",
            "Serial": "Serial", "Firmware": "Firmware", "Status": "Status", "PowerState": "Power State"
        },
    }
    columns_map = {
        "disks": ["Name", "Model", "Serial", "CapacityGB", "Interface", "Type", "Status"],
        "memory": ["Name", "CapacityGB", "Type", "SpeedMHz", "Manufacturer", "Serial", "Status"],
        "cpu": ["Name", "Model", "Manufacturer", "Cores", "Threads", "MaxSpeedMHz", "Status"],
        "network": ["Name", "Description", "MAC", "SpeedMbps", "Status", "Manufacturer", "Model"],
        "power": ["Name", "Model", "Manufacturer", "PowerWatts", "InputVoltage", "Status", "Serial"],
        "fans": ["Name", "Status", "SpeedRPM", "SpeedPercent"],
        "controllers": ["Name", "Model", "Manufacturer", "Firmware", "Status", "Serial", "PartNumber"],
        "systems": ["Name", "Model", "Manufacturer", "Serial", "Firmware", "Status", "PowerState"],
    }

    # Парсим сырые данные в единый формат
    parsers = {
        "disks": parse_disk,
        "memory": parse_memory,
        "cpu": parse_cpu,
        "network": parse_network,
        "power": parse_power,
        "fans": parse_fan,
        "controllers": parse_controller,
        "systems": parse_system,
    }

    for key in selected_types:
        if key not in inventory:
            continue
        raw_data = inventory.get(key, [])
        if not raw_data:
            continue
        parser = parsers.get(key)
        if parser:
            parsed = [parser(item) for item in raw_data]
        else:
            parsed = raw_data
        columns = columns_map.get(key, [])
        field_map = field_maps.get(key, {})
        if columns:
            print_table(key.capitalize(), columns, parsed, field_map)
        else:
            # fallback: вывести как JSON
            console.print(f"[bold]{key.capitalize()}:[/]")
            for item in parsed[:5]:  # ограничим вывод
                console.print(item)
            if len(parsed) > 5:
                console.print(f"... и ещё {len(parsed)-5} записей")


# ============================================
# ЭКСПОРТ В EXCEL
# ============================================

def export_to_excel(inventory: Dict[str, List[Dict]], filename: str, selected_types: List[str]):
    """Экспорт инвентаря в Excel (каждый тип компонентов на отдельном листе)"""
    if not HAS_OPENPYXL:
        console.print("[red]❌ openpyxl не установлен. Установите: pip install openpyxl[/]")
        return

    wb = openpyxl.Workbook()
    # Удаляем стандартный лист
    wb.remove(wb.active)
    parsers = {
        "disks": parse_disk,
        "memory": parse_memory,
        "cpu": parse_cpu,
        "network": parse_network,
        "power": parse_power,
        "fans": parse_fan,
        "controllers": parse_controller,
        "systems": parse_system,
    }

    for key in selected_types:
        raw_data = inventory.get(key, [])
        if not raw_data:
            continue
        parser = parsers.get(key)
        if parser:
            parsed = [parser(item) for item in raw_data]
        else:
            parsed = raw_data
        if not parsed:
            continue
        # Создаём лист
        ws = wb.create_sheet(title=key.capitalize()[:31])  # Excel ограничение на 31 символ
        # Заголовки
        headers = list(parsed[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        # Данные
        for row_idx, item in enumerate(parsed, 2):
            for col_idx, key_header in enumerate(headers, 1):
                val = item.get(key_header)
                ws.cell(row=row_idx, column=col_idx, value=val)
        # Автоширина
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    wb.save(filename)
    console.print(f"[green]✅ Экспортировано в {filename}[/]")


# ============================================
# КОМАНДНАЯ СТРОКА (Argparse)
# ============================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Инвентаризация серверов и СХД через Redfish/Swordfish",
        epilog="Пример: python inventory_cli.py -H 192.168.1.100 -u admin -p pass --disks"
    )
    parser.add_argument("-H", "--host", required=True, help="IP или FQDN устройства (можно с протоколом, например https://192.168.1.1 или http://localhost:8000)")
    parser.add_argument("-u", "--username", required=True, help="Имя пользователя")
    parser.add_argument("-p", "--password", required=True, help="Пароль")
    parser.add_argument("--verify-ssl", action="store_true", help="Включить проверку SSL сертификата (по умолчанию выключена)")
    parser.add_argument("--no-ssl", action="store_true", help="Использовать HTTP вместо HTTPS (для мок-серверов). Игнорируется, если в host указан протокол.")
    parser.add_argument("--disks", action="store_true", help="Показать диски")
    parser.add_argument("--memory", action="store_true", help="Показать память")
    parser.add_argument("--cpu", action="store_true", help="Показать процессоры")
    parser.add_argument("--network", action="store_true", help="Показать сетевые интерфейсы")
    parser.add_argument("--power", action="store_true", help="Показать блоки питания")
    parser.add_argument("--fans", action="store_true", help="Показать вентиляторы")
    parser.add_argument("--controllers", action="store_true", help="Показать контроллеры хранения")
    parser.add_argument("--systems", action="store_true", help="Показать информацию о системах")
    parser.add_argument("--all", action="store_true", help="Показать все компоненты")
    parser.add_argument("--excel", help="Сохранить отчёт в Excel файл (укажите имя файла)")
    parser.add_argument("--json", action="store_true", help="Вывести сырой JSON (для отладки)")
    return parser.parse_args()


def main():
    args = parse_args()

    # Определяем, какие компоненты выводить
    if args.all:
        selected_types = ["disks", "memory", "cpu", "network", "power", "fans", "controllers", "systems"]
    else:
        selected_types = []
        if args.disks:
            selected_types.append("disks")
        if args.memory:
            selected_types.append("memory")
        if args.cpu:
            selected_types.append("cpu")
        if args.network:
            selected_types.append("network")
        if args.power:
            selected_types.append("power")
        if args.fans:
            selected_types.append("fans")
        if args.controllers:
            selected_types.append("controllers")
        if args.systems:
            selected_types.append("systems")
        if not selected_types:
            # Если ничего не выбрано, показываем всё
            selected_types = ["disks", "memory", "cpu", "network", "power", "fans", "controllers", "systems"]

    # Создаем клиент
    client = RedfishClient(
        args.host,
        args.username,
        args.password,
        verify_ssl=args.verify_ssl,
        use_ssl=not args.no_ssl
    )

    console.print(f"[bold blue]🔍 Подключение к {args.host} (SSL: {'выкл' if args.no_ssl else 'вкл'})...[/]")

    try:
        # Сбор инвентаря
        with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"), console=console) as progress:
            task = progress.add_task("Сбор данных...", total=None)
            inventory = collect_inventory(client)
            progress.update(task, completed=True)

        # Если нужен JSON
        if args.json:
            console.print_json(data=inventory)
            return

        # Вывод таблиц
        console.print()
        console.print(Panel.fit(f"[bold green]Инвентарь {args.host}[/]", border_style="green"))
        print_inventory(inventory, selected_types)

        # Экспорт в Excel
        if args.excel:
            export_to_excel(inventory, args.excel, selected_types)

    except httpx.HTTPStatusError as e:
        console.print(f"[red]❌ HTTP ошибка: {e.response.status_code}[/]")
        console.print(e.response.text)
        sys.exit(1)
    except Exception as e:
        console.print(f"[red]❌ Ошибка: {e}[/]")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        client.session.close()


if __name__ == "__main__":
    main()